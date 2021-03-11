__author__ = "Nghia Doan"
__copyright__ = "Copyright 2021"
__version__ = "0.1.0"
__maintainer__ = "Nghia Doan"
__email__ = "nghia71@gmail.com"
__status__ = "Development"


from collections import defaultdict
from datetime import datetime
from json import load, dump, dumps
import re
import sys
import traceback

from openpyxl import load_workbook
from requests import Session

from config_handler import ConfigHandler


config = ConfigHandler('../../conf/harvester.ini')
NLP_URL = config.get_config_option('parr', 'nlp_url')
ALL_HDR = config.get_eval_option('parr', 'hdr_map')
REX_TEL = re.compile('\d{3}(\-|\.)\d{3}(\-|\.)\d{4}(\s+Ext\.\d+)?')


def to_years(value):
    if isinstance(value, int):
        return value

    if '-' in value:
        splits = value.split('-')
        start = int(splits[0])
        end = int('20' + splits[1]) if len(splits[1]) < 4 else int(splits[1])
        return [start, end]
    return int(value)


def to_contact_list(value):
    contact_list = []
    for c in value.split(','):
        contact = dict()
        splits = [s.strip() for s in c.strip().split()]
        if '@' in splits[-1]:
            contact['email'] = splits[-1]
        remain = ' '.join(splits[:-1]).strip()
        match = REX_TEL.search(remain)
        if match:
            contact['tel'] = remain[match.start(): match.end()]
            contact['name'] = remain[:match.start()].strip()
        else:
            contact['name'] = remain
        contact_list.append(contact)
    return contact_list


def to_yn(value):
    if value in ['Yes', 'Y']:
        return True
    elif value in ['Yes', 'Y']:
        return False
    return None


def to_str_list(value):
    if not isinstance(value, str):
        return None
    return [s.strip("and ").strip() for s in value.split(',')]


def to_number(value):
    if isinstance(value, int):
        return int(value)
    elif isinstance(value, float):
        return float(value)
    elif isinstance(value, str):
        try:
            return float(value) if "." in value else int(value)
        except:
            return None
    return None


def transform(case, cell):
    if cell.value in [None, '', 'NA', 'Unknown']:
        return None

    value = cell.value.strip() if isinstance(cell.value, str) else cell.value

    if case == 'number':
        return to_number(value)
    if case == 'integer':
        return int(value) if isinstance(value, int) else None
    elif case == 'year':
        return to_years(value)
    elif case == 'yn':
        return to_yn(value)
    elif case == 'category_list':
        return [s.strip() for s in value.split(',') if s.strip()]
    elif case == 'contact_list':
        return to_contact_list(value)
    elif case in ['ne_org', 'ne_loc']:
        return value if value != 'No specific location available' else None
    elif case == 'ne_org_list':
        return to_str_list(value)
    elif case in ['loc_lat', 'loc_lng']:
        if isinstance(value, float):
            if case == 'loc_lat':
                return value if value >= -90 and value <= 90 else None
            return value
        return None
    return value


def xlsx_to_json(file_name, sheet, data_area):
    session = Session()

    ws = load_workbook(filename=file_name)[sheet]
    hdr_map = {i: [h[1], h[2]] for i, h in enumerate(ALL_HDR)}

    # all_e_set, all_k_set = set(), set()
    # all_v_set = {
    #     h[1]: set() for _, h in enumerate(ALL_HDR)
    #     if h[2] in ['category', 'category_list', 'ne_org', 'ne_org_list', 'ne_loc']
    # }

    row_list, row_no = [], 1
    for row in ws[data_area[0]: data_area[1]]:
        selected = False
        for i, cell in enumerate(row):
            if hdr_map[i][0] != 'selected_project' or cell.value is None:
                continue
            selected = cell.value == 'y'
            print(row_no, hdr_map[i][0], cell.value, selected, len(row_list))
            break
        row_no += 1
        if not selected:
            continue
        # if row[59] != 'y':
        #     continue
        row_map = dict()
        for i, cell in enumerate(row):
            d = transform(hdr_map[i][1], cell)
            if d:
                # row_map[hdr_map[i][0]] = [d, hdr_map[i][1]]
                row_map[hdr_map[i][0]] = d
        row_list.append(row_map)

        # for k, v in row_map.items():
        #     print(k, v)
        #     if k in all_v_set:
        #         if v[1] in ['category_list', 'ne_org_list']:
        #             all_v_set[k].update(v[0])
        #         else:
        #             all_v_set[k].add(v[0])

        # payload = [
        #     {'u': k, 'c': v[0]} for k, v in row_map.items()
        #     if v[1] == 'nlp'
        # ]
        # print('----------')
        # r = session.post(NLP_URL, data=dumps(payload))
        # r.encoding = 'utf-8'
        # for doc in r.json():
        #     for sent in doc['p']:
        #         print(sent['c'])
        #         e_set = set(e['c'] + ' ' + e['t'] for e in sent['e'])
        #         k_set = set(k['c'] for k in sent['k'])
        #         print('\t NE\t', ' | '.join(e_set))
        #         print('\t KP\t', ' | '.join(k_set))
        #         all_e_set.update(e_set)
        #         all_k_set.update(k_set)
        # print('---------------%s---------------' % count)

    # print('------------------------------')
    # for e in sorted(all_e_set):
    #     print(e)
    # print('------------------------------')
    # for k in sorted(all_k_set):
    #     print(k)
    # print('------------------------------')
    # for k, v in all_v_set.items():
    #     print(k, v)

    return row_list


if __name__ == '__main__':

    row_list = xlsx_to_json(sys.argv[1], 'SalmonRestorationActivityList', ['A2', 'BI588'])

    # with open('../data/parr_projects.json', 'wt') as f:
    #     dump(row_list, f)
    # with open('../data/parr_projects.json', 'rt') as f:
    #     o_row_list = load(f)
    #
    # hdr_map = {h[1]: h[2] for h in ALL_HDR}
    # for r1, r2 in zip(row_list, o_row_list):
    #     for k, v in r1.items():
    #         if hdr_map[k] in ['integer', 'number']:
    #             if v and k not in r2 or not r2[k]:
    #                 r2[k] = v
    with open('../../import/parr_projects_2.0.json', 'wt') as f:
        dump(row_list, f)
